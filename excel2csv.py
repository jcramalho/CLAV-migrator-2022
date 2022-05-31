from openpyxl import load_workbook

data_file = '../dados/excel/Frecolha-20220204.xlsx'

# Load the entire workbook.
wb = load_workbook(data_file)

# List all the sheets in the file.
print("Encontrei as seguintes folhas:")
for sheetname in wb.sheetnames:
    print(sheetname)

# Load one worksheet.
ws = wb['100_csv']
all_rows = list(ws.rows)

print(f"Found {len(all_rows)} rows of data.")

print("\nThird row of data:")
for cell in all_rows[2]:
    print(cell.value)