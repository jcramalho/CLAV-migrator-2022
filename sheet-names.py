from openpyxl import load_workbook
wb = load_workbook('../dados/excel/Frecolha-20220204.xlsx')
print(wb.sheetnames)
